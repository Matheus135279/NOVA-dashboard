import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from fpdf import FPDF
import streamlit as st
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import tempfile

def format_currency(value, currency='R$'):
    """Formata valores monetários."""
    try:
        return f"{currency} {float(value):,.2f}"
    except (ValueError, TypeError):
        return f"{currency} 0,00"

def format_number(value, suffix=''):
    """Formata números grandes com K/M/B."""
    try:
        value = float(value)
        if value >= 1_000_000_000:
            return f"{value/1_000_000_000:.1f}B{suffix}"
        elif value >= 1_000_000:
            return f"{value/1_000_000:.1f}M{suffix}"
        elif value >= 1_000:
            return f"{value/1_000:.1f}K{suffix}"
        return f"{value:.0f}{suffix}"
    except (ValueError, TypeError):
        return f"0{suffix}"

def create_evolution_chart(df, metric, title):
    """Cria gráfico de evolução temporal."""
    # Verifica se as colunas necessárias existem
    if 'date' not in df.columns:
        st.error("❌ A coluna 'date' não foi encontrada no arquivo.")
        return None
        
    if metric not in df.columns:
        st.error(f"❌ A coluna '{metric}' não foi encontrada no arquivo.")
        return None
    
    # Seleciona apenas dados numéricos para o gráfico
    if not pd.api.types.is_numeric_dtype(df[metric]):
        st.error(f"❌ A coluna '{metric}' não contém dados numéricos válidos.")
        return None
    
    fig = px.line(
        df,
        x='date',
        y=metric,
        title=title,
        template='plotly_dark'
    )
    
    fig.update_layout(
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font_color='white'
    )
    
    return fig

def create_comparison_chart(df, metric, dimension, title):
    """Cria gráfico de comparação entre dimensões."""
    # Verifica se as colunas necessárias existem
    if dimension not in df.columns:
        st.error(f"❌ A coluna '{dimension}' não foi encontrada no arquivo.")
        return None
        
    if metric not in df.columns:
        st.error(f"❌ A coluna '{metric}' não foi encontrada no arquivo.")
        return None
    
    # Seleciona apenas dados numéricos para o gráfico
    if not pd.api.types.is_numeric_dtype(df[metric]):
        st.error(f"❌ A coluna '{metric}' não contém dados numéricos válidos.")
        return None
    
    # Agrupa os dados
    df_grouped = df.groupby(dimension)[metric].sum().reset_index()
    
    fig = px.bar(
        df_grouped,
        x=dimension,
        y=metric,
        title=title,
        template='plotly_dark'
    )
    
    fig.update_layout(
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        font_color='white'
    )
    
    return fig

def export_to_excel(df, filename):
    """Exporta dados para Excel com formatação."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    
    # Adiciona cabeçalho
    headers = list(df.columns)
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Adiciona dados
    for row, data in enumerate(df.values, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Salva arquivo
    wb.save(filename)

def export_to_pdf(df, charts, filename):
    """Exporta relatório em PDF com dados e gráficos."""
    pdf = FPDF()
    pdf.add_page()
    
    # Título
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Relatório de Performance', 0, 1, 'C')
    pdf.ln(10)
    
    # Dados resumidos
    pdf.set_font('Arial', '', 12)
    for col in df.columns:
        value = df[col].iloc[-1]
        pdf.cell(0, 10, f'{col}: {value}', 0, 1)
    
    # Gráficos
    for chart in charts:
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
            chart.write_image(tmp.name)
            pdf.add_page()
            pdf.image(tmp.name, x=10, y=10, w=190)
            os.unlink(tmp.name)
    
    pdf.output(filename)

def clean_numeric_column(series):
    """
    Converte uma série para números de forma segura, tratando diferentes formatos.
    
    Args:
        series (pd.Series): Série para converter
        
    Returns:
        pd.Series: Série convertida para números
    """
    series = series.astype(str).str.replace('.', '', regex=False)
    series = series.str.replace(',', '.', regex=False)
    series = series.str.extract(r'(\d+\.?\d*)')[0]
    return pd.to_numeric(series, errors='coerce').fillna(0)

def sanitize_dataframe(df):
    """
    Limpa e padroniza tipos de dados no DataFrame para exibição segura no Streamlit.
    
    Args:
        df (pd.DataFrame): DataFrame original
        
    Returns:
        pd.DataFrame: DataFrame limpo e padronizado
    """
    df_clean = df.copy()
    
    # Lista de colunas que devem ser numéricas
    numeric_columns = [
        'spend', 'cost', 'clicks', 'impressions', 'conversions',
        'cpc', 'ctr', 'cpm', 'frequency', 'cost_per_conversion',
        'conversion_value'
    ]
    
    for col in df_clean.columns:
        col_lower = col.lower()
        
        # Trata colunas numéricas conhecidas
        if col_lower in numeric_columns:
            try:
                df_clean[col] = clean_numeric_column(df_clean[col])
            except Exception:
                df_clean[col] = 0
                st.warning(f"⚠️ A coluna '{col}' contém valores inválidos e foi preenchida com zeros.")
        
        # Trata colunas de data
        elif col_lower in ['date', 'data']:
            try:
                df_clean[col] = pd.to_datetime(df_clean[col]).dt.strftime('%d/%m/%Y')
            except Exception:
                df_clean[col] = 'Data inválida'
                st.warning(f"⚠️ A coluna '{col}' contém datas inválidas.")
        
        # Converte outras colunas para string
        else:
            df_clean[col] = df_clean[col].astype(str)
    
    return df_clean

def safe_dataframe_display(df, linhas=5):
    """
    Exibe DataFrame de forma segura no Streamlit, evitando erros de conversão.
    
    Args:
        df (pd.DataFrame): DataFrame para exibir
        linhas (int): Número de linhas a mostrar (default: 5)
    """
    try:
        if df is None or df.empty:
            st.warning("⚠️ Não há dados para exibir.")
            return
            
        df_clean = sanitize_dataframe(df)
        
        # Exibe o DataFrame
        st.dataframe(df_clean.head(linhas))
        
        # Mostra informações úteis
        st.caption(f"Mostrando {min(linhas, len(df))} de {len(df)} linhas. "
                  f"Total de colunas: {len(df.columns)}")
        
    except Exception as e:
        st.error("❌ Erro ao exibir os dados. Verifique se há colunas ou células com valores incompatíveis.")
        if st.checkbox("Mostrar detalhes do erro"):
            st.exception(e)

def clean_for_display(df):
    """Limpa o DataFrame para exibição segura no Streamlit."""
    df_clean = df.copy()
    
    for col in df_clean.columns:
        # Trata valores nulos primeiro
        df_clean[col] = df_clean[col].fillna('N/A')
        
        # Converte datas para string no formato brasileiro
        if pd.api.types.is_datetime64_any_dtype(df_clean[col]):
            df_clean[col] = df_clean[col].dt.strftime('%d/%m/%Y')
        
        # Formata números com 2 casas decimais
        elif pd.api.types.is_float_dtype(df_clean[col]):
            df_clean[col] = df_clean[col].apply(lambda x: f"{x:,.2f}" if pd.notnull(x) else 'N/A')
        
        # Formata números inteiros sem decimais
        elif pd.api.types.is_integer_dtype(df_clean[col]):
            df_clean[col] = df_clean[col].apply(lambda x: f"{x:,}" if pd.notnull(x) else 'N/A')
        
        # Converte categorias para string
        elif df_clean[col].dtype.name == 'category':
            df_clean[col] = df_clean[col].astype(str)
        
        # Converte outros tipos para string
        else:
            df_clean[col] = df_clean[col].astype(str)
    
    return df_clean

def map_csv_columns(df):
    """Mapeia colunas do CSV para nomes padronizados e converte tipos."""
    column_mapping = {
        # Mapeamento Facebook Ads
        "nome da campanha": "campaign",
        "nome do conjunto de anúncios": "campaign",
        "campanha": "campaign",
        "valor usado (brl)": "cost",
        "custo": "cost",
        "valor gasto": "cost",
        "dia": "date",
        "data": "date",
        "data do relatório": "date",
        "cliques no link": "clicks",
        "cliques": "clicks",
        "cliques totais": "clicks",
        "cpc (custo por clique no link)": "cpc",
        "cpc": "cpc",
        "custo por clique": "cpc",
        "ctr (taxa de cliques no link)": "ctr",
        "ctr": "ctr",
        "taxa de cliques": "ctr",
        "resultados": "conversions",
        "conversões": "conversions",
        "ações": "conversions",
        "valor de conversão": "conversion_value",
        "valor das conversões": "conversion_value",
        "retorno": "conversion_value",
        "impressões": "impressions",
        "visualizações": "impressions",
        "alcance": "impressions",
        "frequência": "frequency",
        "cpm (custo por 1.000 impressões)": "cpm",
        "objetivo": "objective",
        "veiculação da campanha": "campaign_delivery",
        "orçamento da campanha": "campaign_budget",
        "tipo de orçamento da campanha": "campaign_budget_type",
        "tipo de resultado": "conversion_type",
        "custo por resultado": "cost_per_conversion"
    }
    
    # Tenta mapear cada coluna
    mapped_columns = {}
    missing_columns = []
    
    for col in df.columns:
        col_lower = col.lower().strip()
        if col_lower in column_mapping:
            mapped_columns[col] = column_mapping[col_lower]
        else:
            missing_columns.append(col)
    
    # Se encontrou colunas não mapeadas, exibe aviso
    if missing_columns:
        st.warning(f"⚠️ As seguintes colunas não foram mapeadas e serão mantidas como estão: {', '.join(missing_columns)}")
    
    # Aplica o mapeamento
    df_mapped = df.rename(columns=mapped_columns)
    
    # Limpa e padroniza o DataFrame
    df_mapped = sanitize_dataframe(df_mapped)
    
    return df_mapped

def calculate_kpis(df):
    """Calcula KPIs principais."""
    kpis = {}
    
    # Verifica se temos a coluna campaign
    if 'campaign' not in df.columns:
        st.error("❌ A coluna 'campaign' não foi encontrada no arquivo.")
        return kpis
    
    # Seleciona apenas colunas numéricas para agregação
    numeric_columns = df.select_dtypes(include=['int64', 'float64']).columns
    metrics_to_sum = [col for col in numeric_columns if col != 'date']
    
    # Agrupa por campanha apenas as métricas numéricas
    if metrics_to_sum:
        df_grouped = df.groupby('campaign')[metrics_to_sum].sum().reset_index()
    else:
        st.warning("⚠️ Nenhuma coluna numérica encontrada para agregação.")
        return kpis
    
    # Calcula KPIs com verificação de existência das colunas
    if 'impressions' in df_grouped.columns:
        kpis['Impressões'] = df_grouped['impressions'].sum()
    else:
        kpis['Impressões'] = 0
        
    if 'clicks' in df_grouped.columns:
        kpis['Cliques'] = df_grouped['clicks'].sum()
    else:
        kpis['Cliques'] = 0
        
    if all(col in df_grouped.columns for col in ['clicks', 'impressions']) and df_grouped['impressions'].sum() > 0:
        kpis['CTR'] = (df_grouped['clicks'].sum() / df_grouped['impressions'].sum() * 100)
    else:
        kpis['CTR'] = 0
        
    if all(col in df_grouped.columns for col in ['cost', 'clicks']) and df_grouped['clicks'].sum() > 0:
        kpis['CPC Médio'] = df_grouped['cost'].sum() / df_grouped['clicks'].sum()
    else:
        kpis['CPC Médio'] = 0
        
    if 'conversions' in df_grouped.columns:
        kpis['Conversões'] = df_grouped['conversions'].sum()
    else:
        kpis['Conversões'] = 0
        
    if 'cost' in df_grouped.columns:
        kpis['Custo Total'] = df_grouped['cost'].sum()
    else:
        kpis['Custo Total'] = 0
        
    if all(col in df_grouped.columns for col in ['conversion_value', 'cost']) and df_grouped['cost'].sum() > 0:
        kpis['ROAS'] = df_grouped['conversion_value'].sum() / df_grouped['cost'].sum()
    else:
        kpis['ROAS'] = 0
        
    # Adiciona métricas adicionais se disponíveis
    if 'frequency' in df_grouped.columns:
        kpis['Frequência Média'] = df_grouped['frequency'].mean()
        
    if 'cpm' in df_grouped.columns:
        kpis['CPM Médio'] = df_grouped['cpm'].mean()
        
    if 'cost_per_conversion' in df_grouped.columns and df_grouped['cost_per_conversion'].sum() > 0:
        kpis['Custo por Conversão Médio'] = df_grouped['cost_per_conversion'].mean()
    
    return kpis 